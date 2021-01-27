from flask import Flask, render_template, request, jsonify
from pyhive import hive
import urllib
from io import BytesIO
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from flask import make_response

app = Flask(__name__)

solr_host = "192.168.17.1:8983"

#@app.route("/")
#def home():
#    return render_template("index.html")

# routes
@app.route("/", methods=["GET", "POST"])
def home2():
	if(request.method == "GET"):
		return render_template("index2.html")
	else:
		fields = {}
		data = request.form.to_dict()
		mssv = data.pop('mssv', '')
		for key in data.keys():
			if(key == 'thongtin'):
				fields[key] = du_lieu_sinh_vien(mssv)
			if(key == 'diemthi'):
				fields[key] = du_lieu_diem_thi(mssv)
			if(key == 'diemdanh'):
				fields[key] = du_lieu_diem_danh(mssv)
			if(key == 'khenthuong'):
				fields[key] = du_lieu_khen_thuong(mssv)
			if(key == 'kyluat'):
				fields[key] = du_lieu_thoi_hoc(mssv)
			if(key == 'hocphi'):
				fields[key] = du_lieu_hoc_phi(mssv)

		return render_template("index2.html", mssv=mssv, dulieu=list(data.keys()), **fields)

@app.route("/xuat-file")
def xuat_file():
	data = {}
	params = request.args.to_dict(flat=False)
	mssv = params.get('mssv')[0]
	for field in params.get('dulieu'):
		if field == 'thongtin':
			data[field] = du_lieu_sinh_vien(mssv)
		if field == 'diemthi':
			data[field] = du_lieu_diem_thi(mssv)
		if field == 'khenthuong':
			data[field] = du_lieu_khen_thuong(mssv)
		if field == 'kyluat':
			data[field] = du_lieu_thoi_hoc(mssv)
		if field == 'hocphi':
			data[field] = du_lieu_hoc_phi(mssv)
	wb = Workbook()
	mainsheet = wb.create_sheet("Sinh vien")
	for k, v in data.items():
		if k == 'thongtin':
			if v != None:
				ws = mainsheet
				ws['A1'] = 'Ma so sinh vien'
				ws['B1'] = v[0]
				ws['A2'] = 'Lop'
				ws['B2'] = v[1]
				ws['A3'] = 'Ho va ten'
				ws['B3'] = v[2]
				ws['A4'] = 'PortalID'
				ws['B4'] = v[7]
				ws['A5'] = 'Gioi tinh'
				ws['B5'] = "Nam" if v[6] else "Nu"
				ws['A6'] = 'Ngay sinh'
				ws['B6'] = str(v[3]) + '/' + str(v[4]) + '/' + str(v[5]) if(v[3] and v[4] and v[5]) else ''
				ws['A7'] = 'Dien thoai'
				ws['B7'] = v[8]
				ws['A8'] = 'Zalo'
				ws['B8'] = v[20]
				ws['A9'] = 'Email'
				ws['B9'] = v[10]
				ws['A10'] = 'Dia chi'
				ws['B10'] = v[9]
				ws['A11'] = 'CMND'
				ws['B11'] = v[15]
				ws['A12'] = 'Ngay cap'
				ws['B12'] = v[17]
				ws['A13'] = 'Dia chi'
				ws['B13'] = v[16]
				ws['A14'] = 'Ho ten nguoi than'
				ws['B14'] = v[12]
				ws['A15'] = 'Moi lien he'
				ws['B15'] = v[18]
				ws['A16'] = 'Email nguoi than'
				ws['B16'] = v[13]
				ws['A17'] = 'So dien thoai nguoi than'
				ws['B17'] = v[14]
				ws['A18'] = 'Zalo nguoi than'
				ws['B18'] = v[19]
				ws['A19'] = 'Ghi chu'
				ws['B19'] = v[11]

		if k == 'khenthuong':
			if v != None:
				mainsheet['A20'] = 'Khen thuong'
				mainsheet['B20'] = v

		if k == 'kyluat':
			if v != None:
				mainsheet['A21'] = 'Ky luat'
				mainsheet['B21'] = v
			
		if k == 'diemthi':
			if v != None:
				ws = wb.create_sheet("Diem thi")
				ws['A1'] = 'Mon hoc'
				ws['B1'] = 'Loai thi'
				ws['C1'] = 'Ngay thi'
				ws['D1'] = 'Diem thi'
				ws['E1'] = 'Lan thi'
				for i in range(len(v)):
					for j in range(5):
						ws.cell(row=2+i, column=1+j, value=v[i][j])

		if k == 'hocphi':
			if v != None:
				ws = wb.create_sheet("Hoc phi")
				ws['A1'] = 'Ngay dong'
				ws['B1'] = 'So tien'
				for i in range(len(v)):
					ws.cell(row=2+i, column=1, value=v[i][0])
					ws.cell(row=2+i, column=2, value=v[i][1])

		if k == 'diemdanh':
			if v != None:
				ws = wb.create_sheet('Diem danh')
				for i in range(12):
					ws.cell(row=2, column=i, value=v[i+3])

	content = save_virtual_workbook(wb)
	resp = make_response(content)
	resp.headers["Content-Disposition"] = 'attachment; filename='+mssv+'.xlsx'
	resp.headers['Content-Type'] = 'application/x-xlsx'
	return resp


# Search from Solr
def du_lieu_khen_thuong(mssv):
	print("http://"+solr_host+"/solr/khen_thuong/select?q=_text%3A"+mssv+"&wt=python")
	connection = urllib.request.urlopen("http://"+solr_host+"/solr/khen_thuong/select?q=_text%3A"+mssv+"&wt=python")
	response = eval(connection.read()) # dictionary
	if(response["response"]["numFound"] == 0):
		return None
	return response["response"]["docs"][0]
	
def du_lieu_thoi_hoc(mssv):
	connection = urllib.request.urlopen("http://"+solr_host+"/solr/thoi_hoc/select?q=_text%3A"+mssv+"&wt=python")
	response = eval(connection.read()) # dictionary
	if(response["response"]["numFound"] == 0):
		return None
	return response["response"]["docs"][0] # dictionary

# Search from Cloudera	
def du_lieu_diem_danh(mssv):
	conn = hive.connect(host='quickstart.cloudera', port = 10000)
	cursor = conn.cursor()
	cursor.execute("""SELECT * FROM `default`.`ds_diem_danh` WHERE `?roll number`=%s""",(mssv,))
	result = []
	for row in cursor.fetchall():
		result.append(row)
	return result[0] if len(result) != 0 else None

def du_lieu_sinh_vien(mssv):
	conn = hive.connect(host='quickstart.cloudera', port = 10000)
	cursor = conn.cursor()
	cursor.execute("""SELECT *from aptech_dmsinhvien where sv_mssv=%s""",(mssv,))
	result = []
	for row in cursor.fetchall():
		result.append(row)
	return result[0] if len(result) != 0 else None

def du_lieu_diem_thi(mssv):
	conn = hive.connect(host='quickstart.cloudera', port = 10000)
	cursor = conn.cursor()
	cursor.execute("""SELECT mh.MH_TEN, lt.LOAITHI_TEN, dm.KT_NGAY, tcd.THI_DIEM, dm.KT_LANTHI FROM aptech_tcdthi tcd join aptech_dmkythi dm on tcd.KT_ID = dm.KT_ID JOIN aptech_dm_loaithi lt on lt.LOAITHI_ID = dm.KT_LOAITHI join aptech_dmmonhoc mh on dm.MH_ID = mh.MH_ID WHERE tcd.sv_mssv=%s""",(mssv,))
	result = []
	for row in cursor.fetchall():
		result.append(row)
	return result if len(result) != 0 else None

def du_lieu_hoc_phi(mssv):
	conn = hive.connect(host='quickstart.cloudera', port = 10000)
	cursor = conn.cursor()
	cursor.execute("""SELECT *from hocphi where mssv=%s""",(mssv,))
	result = []
	for row in cursor.fetchall():
		result.append(row)
	return result if len(result) != 0 else None


if __name__ == "__main__":
    app.run(host='0.0.0.0', debug=True)
